#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
WPS Excel修复工具 - 一体化版本
包含所有必需的功能模块
"""

# ========== 核心修复模块 ==========
import zipfile
import os
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
import io
from PIL import Image as PILImage


class PreciseSafeWPSExcelFixer:
    """精确且安全的WPS Excel修复工具，结合perfect.py的精确计算和safe.py的安全特性"""
    
    def __init__(self, xlsx_file_path):
        self.xlsx_file_path = xlsx_file_path
        self.image_list = []
        self.workbook = None
        
    def analyze_dispimg_cells(self):
        """分析所有工作表中包含DISPIMG公式的单元格"""
        try:
            self.workbook = openpyxl.load_workbook(self.xlsx_file_path, data_only=False)
        except Exception as e:
            print(f"无法加载Excel文件: {e}")
            return {}
            
        all_dispimg_cells = {}
        
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            dispimg_cells = []
            
            try:
                for row in sheet.iter_rows(min_row=1, values_only=False):
                    for cell in row:
                        if (cell.value and isinstance(cell.value, str) and 
                            '=_xlfn.DISPIMG(' in cell.value):
                            formula = cell.value
                            start = formula.find('"') + 1
                            end = formula.find('"', start)
                            image_id = formula[start:end]
                            
                            cell_info = {
                                'cell': cell,
                                'sheet': sheet,
                                'sheet_name': sheet_name,
                                'row': cell.row,
                                'column': cell.column,
                                'image_id': image_id,
                                'coordinate': cell.coordinate
                            }
                            dispimg_cells.append(cell_info)
                            self.image_list.append(image_id)
            except Exception as e:
                print(f"分析工作表 {sheet_name} 时出错: {e}")
                continue
                
            if dispimg_cells:
                all_dispimg_cells[sheet_name] = dispimg_cells
                print(f"  发现 {len(dispimg_cells)} 个DISPIMG公式")
                    
        return all_dispimg_cells
    
    def get_image_mapping(self):
        """获取图片ID到文件路径的映射关系"""
        try:
            with zipfile.ZipFile(self.xlsx_file_path, 'r') as zfile:
                required_files = ['xl/cellimages.xml', 'xl/_rels/cellimages.xml.rels']
                for req_file in required_files:
                    if req_file not in zfile.namelist():
                        print(f"缺少必要文件: {req_file}")
                        return {}
                        
                xml_content = zfile.read('xl/cellimages.xml')
                relxml_content = zfile.read('xl/_rels/cellimages.xml.rels')
        except Exception as e:
            print(f"读取图片映射时出错: {e}")
            return {}

        try:
            root = ET.fromstring(xml_content)
            name_to_embed_map = {}
            
            namespaces = {
                'etc': 'http://www.wps.cn/officeDocument/2017/etCustomData',
                'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'
            }
            
            for cell_image in root.findall('.//etc:cellImage', namespaces):
                name_elem = cell_image.find('.//xdr:cNvPr', namespaces)
                embed_elem = cell_image.find('.//a:blip', namespaces)
                
                if name_elem is not None and embed_elem is not None:
                    name = name_elem.attrib['name']
                    embed = embed_elem.attrib['{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed']
                    name_to_embed_map[name] = embed

            root1 = ET.fromstring(relxml_content)
            namespaces = {'r': 'http://schemas.openxmlformats.org/package/2006/relationships'}
            
            id_target_map = {
                child.attrib['Id']: child.attrib.get('Target', 'No Target Found') 
                for child in root1.findall('.//r:Relationship', namespaces=namespaces)
            }
            
            name_to_target_map = {
                name: id_target_map[embed] 
                for name, embed in name_to_embed_map.items() 
                if embed in id_target_map
            }
            
            return name_to_target_map
            
        except Exception as e:
            print(f"解析XML时出错: {e}")
            return {}
    
    def extract_image_from_xlsx(self, image_path):
        """从xlsx文件中提取指定路径的图片数据"""
        try:
            with zipfile.ZipFile(self.xlsx_file_path, 'r') as zfile:
                actual_image_path = f'xl/{image_path}'
                if actual_image_path in zfile.namelist():
                    return zfile.read(actual_image_path)
        except Exception as e:
            print(f"提取图片数据时出错: {e}")
        return None
    
    def get_precise_cell_dimensions(self, sheet, cell_info):
        """精确计算单元格尺寸，优化缩放算法"""
        try:
            row = cell_info['row']
            column = cell_info['column']
            coordinate = cell_info['coordinate']
            sheet_name = cell_info['sheet_name']
            
            column_letter = get_column_letter(column)
            
            # 安全获取列宽和行高
            try:
                column_width = sheet.column_dimensions[column_letter].width or 8.43
            except (AttributeError, KeyError):
                column_width = 8.43
                
            try:
                row_height = sheet.row_dimensions[row].height or 15
            except (AttributeError, KeyError):
                row_height = 15
            
            # 确保数值有效
            column_width = float(column_width) if column_width else 8.43
            row_height = float(row_height) if row_height else 15
            
            # 优化像素转换算法
            # 增加基础倍数，让图片更大更清晰
            base_width_multiplier = 12.0  # 从7.5增加到12
            base_height_multiplier = 6.0   # 从4.5增加到6
            
            cell_width_px = max(int(column_width * base_width_multiplier), 100)   # 最小100px
            cell_height_px = max(int(row_height * base_height_multiplier), 80)    # 最小80px
            
            calc_type = "优化标准计算"
            
            # 合理范围控制
            cell_width_px = max(100, min(cell_width_px, 800))   # 扩大最大尺寸
            cell_height_px = max(80, min(cell_height_px, 600))  # 扩大最大尺寸
            
            print(f"  单元格 {coordinate} [{calc_type}]: {cell_width_px}x{cell_height_px} 像素")
            
            return cell_width_px, cell_height_px
            
        except Exception as e:
            print(f"  获取单元格尺寸时出错 {cell_info['coordinate']}: {e}")
            return 150, 120  # 更大的安全默认值
    
    def calculate_proper_scaling(self, cell_width_px, cell_height_px, image_width, image_height):
        """优化的等比例缩放算法，生成更大更清晰的图片"""
        if image_width <= 0 or image_height <= 0:
            return 120, 90  # 增大默认值
        
        # 保持原始比例
        original_ratio = image_width / image_height
        
        # 更激进的缩放策略 - 允许图片占用更多空间
        max_width = cell_width_px * 1.2   # 从0.9增加到1.2，允许超出单元格
        max_height = cell_height_px * 1.2  # 从0.9增加到1.2
        
        # 计算等比例缩放
        width_scale = max_width / image_width
        height_scale = max_height / image_height
        scale_factor = min(width_scale, height_scale)
        
        # 增加缩放系数下限，确保图片不会太小
        scale_factor = max(scale_factor, 0.08)  # 最小缩放比例8%
        
        # 计算最终尺寸，设置更大的最小值
        final_width = max(int(image_width * scale_factor), 60)   # 从25增加到60
        final_height = max(int(image_height * scale_factor), 45)  # 从25增加到45
        
        # 确保比例正确
        calculated_ratio = final_width / final_height
        if abs(calculated_ratio - original_ratio) > 0.01:
            if original_ratio > 1:  # 宽图
                final_height = max(int(final_width / original_ratio), 45)
            else:  # 高图或方图
                final_width = max(int(final_height * original_ratio), 60)
        
        print(f"  缩放计算: 原始{image_width}x{image_height} -> 单元格{cell_width_px}x{cell_height_px} -> 最终{final_width}x{final_height}")
        
        return final_width, final_height
    
    def create_safe_anchor(self, cell_info, final_width_px, final_height_px):
        """创建精确的图片锚点，直接锚定到原始单元格"""
        try:
            from openpyxl.drawing.xdr import XDRPositiveSize2D
            
            # 使用原始单元格位置（0-based索引）
            row = cell_info['row'] - 1
            column = cell_info['column'] - 1
            
            # 转换为EMU单位
            final_width_emu = pixels_to_EMU(final_width_px)
            final_height_emu = pixels_to_EMU(final_height_px)
            
            # 创建锚点：直接定位到原始单元格的左上角，无额外偏移
            marker = AnchorMarker(col=column, colOff=0, row=row, rowOff=0)
            size = XDRPositiveSize2D(cx=final_width_emu, cy=final_height_emu)
            anchor = OneCellAnchor(_from=marker, ext=size)
            
            print(f"    创建锚点: 单元格({row}, {column}) -> 图片{final_width_px}x{final_height_px}")
            
            return anchor
            
        except Exception as e:
            print(f"    创建锚点时出错: {e}")
            return None
    
    def fix_excel_file_precise_safe(self, output_path=None):
        """精确且安全的修复Excel文件"""
        if output_path is None:
            output_path = self.xlsx_file_path.replace('.xlsx', '_fixed.xlsx')  # 统一使用_fixed后缀
        
        print("=== 精确安全WPS图片修复工具 ===")
        print("正在分析所有工作表中的DISPIMG单元格...")
        
        all_dispimg_cells = self.analyze_dispimg_cells()
        if not all_dispimg_cells:
            print("未发现需要修复的DISPIMG公式")
            return
            
        total_cells = sum(len(cells) for cells in all_dispimg_cells.values())
        print(f"总共发现 {total_cells} 个DISPIMG公式需要修复")
        
        print("正在获取图片映射关系...")
        image_mapping = self.get_image_mapping()
        if not image_mapping:
            print("未找到图片映射关系")
            return
        
        successful_fixes = 0
        
        for sheet_name, dispimg_cells in all_dispimg_cells.items():
            print(f"\n正在处理工作表: {sheet_name}")
            sheet = self.workbook[sheet_name]
            
            for cell_info in dispimg_cells:
                image_id = cell_info['image_id']
                cell = cell_info['cell']
                
                if image_id in image_mapping:
                    image_path = image_mapping[image_id]
                    print(f"\n  正在处理图片: {image_id}")
                    
                    image_data = self.extract_image_from_xlsx(image_path)
                    if image_data:
                        try:
                            # 获取图片原始尺寸
                            img_stream = io.BytesIO(image_data)
                            with PILImage.open(img_stream) as pil_img:
                                original_width, original_height = pil_img.size
                            
                            # 获取精确单元格尺寸
                            cell_width_px, cell_height_px = self.get_precise_cell_dimensions(sheet, cell_info)
                            
                            # 计算精确缩放
                            final_width_px, final_height_px = self.calculate_proper_scaling(
                                cell_width_px, cell_height_px, original_width, original_height)
                            
                            # 清除原始公式
                            cell.value = None
                            
                            # 创建图片对象
                            img_stream.seek(0)
                            img = OpenpyxlImage(img_stream)
                            img.width = final_width_px
                            img.height = final_height_px
                            
                            # 创建安全锚点
                            anchor = self.create_safe_anchor(cell_info, final_width_px, final_height_px)
                            if anchor:
                                img.anchor = anchor
                                sheet.add_image(img)
                                
                                successful_fixes += 1
                                print(f"  [OK] 成功修复: {cell.coordinate} -> {final_width_px}x{final_height_px}")
                            else:
                                print(f"  [FAIL] 锚点创建失败: {cell.coordinate}")
                                
                        except Exception as e:
                            print(f"  [FAIL] 修复失败: {cell.coordinate} - {str(e)}")
                            cell.value = f'=_xlfn.DISPIMG("{image_id}")'
                    else:
                        print(f"  [FAIL] 无法提取图片数据: {image_id}")
                else:
                    print(f"  [FAIL] 未找到图片映射: {image_id}")
        
        # 清理兼容性设置
        print("\n正在清理兼容性设置...")
        try:
            for sheet in self.workbook.worksheets:
                if hasattr(sheet, '_ext_lst'):
                    sheet._ext_lst = None
        except Exception as e:
            print(f"清理兼容性设置时出错: {e}")
        
        # 保存文件
        print(f"正在保存修复后的文件到: {output_path}")
        try:
            self.workbook.save(output_path)
            print(f"\n[COMPLETE] 修复完成!")
            print(f"总计处理: {total_cells} 个DISPIMG公式")
            print(f"成功修复: {successful_fixes} 个")
            print(f"输出文件: {output_path}")
            return output_path
            
        except Exception as e:
            print(f"保存文件时出错: {e}")
            return None
    
    def preview_fixes(self):
        """预览将要修复的内容"""
        print("=== 预览修复内容 ===")
        all_dispimg_cells = self.analyze_dispimg_cells()
        
        if not all_dispimg_cells:
            print("未发现需要修复的内容")
            return
            
        image_mapping = self.get_image_mapping()
        
        total = 0
        for sheet_name, cells in all_dispimg_cells.items():
            print(f"\n工作表 '{sheet_name}':")
            for cell_info in cells:
                image_id = cell_info['image_id']
                if image_id in image_mapping:
                    print(f"  {cell_info['coordinate']}: {image_id} -> {image_mapping[image_id]}")
                else:
                    print(f"  {cell_info['coordinate']}: {image_id} (图片映射缺失)")
                total += 1
        
        print(f"\n总计发现 {total} 个需要修复的图片")


def main():
    """主函数 - 仅供单独运行时使用"""
    # 当直接运行此文件时才执行修复
    # GUI模式下不执行，避免重复调用
    pass


if __name__ == '__main__':
    main()

# ========== GUI主程序 ==========
import tkinter as tk
from tkinter import ttk, messagebox
import os
import threading
import time
import sys
from pathlib import Path

# 直接使用内嵌的修复类，无需导入
# # 使用内嵌的PreciseSafeWPSExcelFixer类

class ProgressWindow:
    """现代化进度窗口类"""
    def __init__(self, file_path):
        self.file_path = file_path
        self.repaired_file = None
        
        # 创建进度窗口
        self.window = tk.Tk()
        self.window.title("WPS Excel 图片修复工具")
        self.window.geometry("400x150")
        self.window.resizable(False, False)
        
        # 设置窗口透明度和现代化外观
        self.window.attributes('-alpha', 0.95)
        try:
            self.window.attributes('-topmost', False)
        except:
            pass
        
        # 设置窗口图标
        try:
            icon_loaded = False
            
            # 方法1：在打包环境中尝试从资源路径加载
            if hasattr(sys, '_MEIPASS'):
                icon_path = os.path.join(sys._MEIPASS, 'assset', 'icon.ico')
                if os.path.exists(icon_path):
                    self.window.iconbitmap(icon_path)
                    icon_loaded = True
                    print(f"已从打包资源加载图标: {icon_path}")
            
            # 方法2：在开发环境中从相对路径加载
            if not icon_loaded:
                icon_path = os.path.join(os.path.dirname(__file__), 'assset', 'icon.ico')
                if os.path.exists(icon_path):
                    self.window.iconbitmap(icon_path)
                    icon_loaded = True
                    print(f"已从相对路径加载图标: {icon_path}")
            
            # 方法3：从当前目录尝试加载
            if not icon_loaded:
                icon_path = os.path.join(os.getcwd(), 'assset', 'icon.ico')
                if os.path.exists(icon_path):
                    self.window.iconbitmap(icon_path)
                    icon_loaded = True
                    print(f"已从当前目录加载图标: {icon_path}")
            
            if not icon_loaded:
                print("警告: 无法加载自定义图标，使用系统默认图标")
                
        except Exception as e:
            print(f"加载图标失败: {e}")
        
        # 居中显示窗口
        self.center_window()
        
        # 配置样式
        self.setup_styles()
        self.create_widgets()
        
        # 启动修复
        self.start_repair()
    
    def center_window(self):
        """将窗口居中显示"""
        self.window.update_idletasks()
        screen_width = self.window.winfo_screenwidth()
        screen_height = self.window.winfo_screenheight()
        
        x = (screen_width // 2) - (400 // 2)
        y = (screen_height // 2) - (150 // 2)
        
        self.window.geometry(f"400x150+{x}+{y}")
    
    def setup_styles(self):
        """设置现代化样式"""
        self.colors = {
            'primary': '#2563EB',      # 现代蓝色
            'secondary': '#8B5CF6',    # 紫色渐变
            'success': '#10B981',      # 绿色
            'warning': '#F59E0B',      # 橙色
            'error': '#EF4444',        # 红色
            'background': '#FFFFFF',   # 纯白背景
            'card': '#F8FAFC',         # 卡片背景
            'border': '#E2E8F0',       # 边框色
            'text_primary': '#1E293B', # 主文本
            'text_secondary': '#64748B', # 次要文本
            'shadow': '#E0E0E0'        # 阴影颜色（移除透明度）
        }
        
        # 创建简洁背景（移除渐变）
        self.window.configure(bg=self.colors['background'])
        
        # 配置ttk样式
        style = ttk.Style()
        style.theme_use('clam')
        
        # 配置框架样式
        style.configure('Card.TFrame', 
                       background=self.colors['card'],
                       relief='flat',
                       borderwidth=1)
        
        # 配置圆角无边框进度条样式
        style.configure('Rounded.Horizontal.TProgressbar',
                       background=self.colors['primary'],
                       troughcolor=self.colors['border'],
                       borderwidth=0,
                       relief='flat',
                       thickness=12,
                       focuscolor='none',
                       lightcolor=self.colors['primary'],
                       darkcolor=self.colors['primary'],
                       arrowcolor=self.colors['primary'])
        
        # 配置进度条的圆角样式
        style.map('Rounded.Horizontal.TProgressbar',
                  background=[('active', self.colors['primary']),
                             ('!active', self.colors['primary'])],
                  relief=[('active', 'flat'),
                         ('!active', 'flat')],
                  borderwidth=[('active', '0'),
                              ('!active', '0')])
        
        # 配置标签样式
        style.configure('Title.TLabel',
                       background=self.colors['background'],
                       foreground=self.colors['text_primary'],
                       font=('Microsoft YaHei UI', 16, 'bold'))
        
        style.configure('Subtitle.TLabel',
                       background=self.colors['background'],
                       foreground=self.colors['text_secondary'],
                       font=('Microsoft YaHei UI', 10))
        
        style.configure('Status.TLabel',
                       background=self.colors['card'],
                       foreground=self.colors['text_primary'],
                       font=('Microsoft YaHei UI', 11))
    
    def create_widgets(self):
        """创建简化窗口组件"""
        # 主容器
        main_container = tk.Frame(self.window, bg=self.colors['background'])
        main_container.pack(fill='both', expand=True, padx=20, pady=20)
        
        # 百分比显示（在进度条上方）
        self.progress_percent_label = tk.Label(main_container, text="0%",
                                             font=('Microsoft YaHei UI', 20, 'bold'),
                                             fg=self.colors['text_primary'], 
                                             bg=self.colors['background'])
        self.progress_percent_label.pack(pady=(0, 10))
        
        # 进度条 - 使用Canvas绘制自定义圆角进度条
        self.progress_var = tk.DoubleVar()
        self.progress_canvas = tk.Canvas(main_container, 
                                       width=360, height=12,
                                       bg=self.colors['background'],
                                       highlightthickness=0,
                                       relief='flat',
                                       borderwidth=0)
        self.progress_canvas.pack(pady=(0, 15))
        
        # 绘制初始进度条背景
        self.draw_progress_bar(0)
        
        # 状态标签（在进度条下方）
        self.status_label = tk.Label(main_container, text="准备开始...",
                                   font=('Microsoft YaHei UI', 10),
                                   fg=self.colors['text_secondary'], 
                                   bg=self.colors['background'])
        self.status_label.pack()
    
    def draw_progress_bar(self, progress):
        """绘制自定义圆角进度条"""
        self.progress_canvas.delete("all")
        
        # 进度条参数
        width = 360
        height = 12
        radius = 6  # 圆角半径
        
        # 背景色（灰色轨道）
        bg_color = self.colors['border']
        
        # 进度色
        if progress >= 100:
            fg_color = self.colors['success']
        elif progress >= 80:
            fg_color = self.colors['warning'] 
        else:
            fg_color = self.colors['primary']
        
        # 绘制背景圆角矩形
        self.draw_rounded_rectangle(self.progress_canvas, 0, 0, width, height, radius, bg_color)
        
        # 绘制进度圆角矩形
        if progress > 0:
            progress_width = max(radius * 2, int(width * progress / 100))  # 确保至少显示圆角
            self.draw_rounded_rectangle(self.progress_canvas, 0, 0, progress_width, height, radius, fg_color)
    
    def draw_rounded_rectangle(self, canvas, x1, y1, x2, y2, radius, fill_color):
        """在Canvas上绘制圆角矩形"""
        # 绘制中间的矩形
        canvas.create_rectangle(x1 + radius, y1, x2 - radius, y2, fill=fill_color, outline="")
        canvas.create_rectangle(x1, y1 + radius, x2, y2 - radius, fill=fill_color, outline="")
        
        # 绘制四个圆角
        canvas.create_oval(x1, y1, x1 + radius * 2, y1 + radius * 2, fill=fill_color, outline="")  # 左上
        canvas.create_oval(x2 - radius * 2, y1, x2, y1 + radius * 2, fill=fill_color, outline="")  # 右上
        canvas.create_oval(x1, y2 - radius * 2, x1 + radius * 2, y2, fill=fill_color, outline="")  # 左下
        canvas.create_oval(x2 - radius * 2, y2 - radius * 2, x2, y2, fill=fill_color, outline="")  # 右下
    
    def update_status(self, message, detail=""):
        """更新状态"""
        self.status_label.config(text=message)
        self.window.update()
    
    def update_progress(self, value, message=""):
        """更新进度"""
        self.progress_percent_label.config(text=f"{int(value)}%")
        
        # 绘制进度条
        self.draw_progress_bar(value)
        
        # 根据进度改变百分比颜色
        if value >= 100:
            self.progress_percent_label.config(fg=self.colors['success'])
        elif value >= 80:
            self.progress_percent_label.config(fg=self.colors['warning'])
        else:
            self.progress_percent_label.config(fg=self.colors['primary'])
        
        if message:
            self.status_label.config(text=message)
        self.window.update()
    
    def start_repair(self):
        """开始修复"""
        repair_thread = threading.Thread(target=self.repair_worker)
        repair_thread.daemon = True
        repair_thread.start()
    
    def repair_worker(self):
        """修复工作线程"""
        try:
            self.update_status("正在加载文件...", "开始修复过程...")
            
            # 创建修复器
            fixer = PreciseSafeWPSExcelFixer(self.file_path)
            
            # 设置输出路径 - 确保文件名统一
            output_path = self.file_path.replace('.xlsx', '_fixed.xlsx')
            
            # 直接调用核心修复方法
            self.update_progress(10, "正在执行修复...")
            
            try:
                # 只调用一次修复方法，避免重复
                result = fixer.fix_excel_file_precise_safe(output_path)
                
                if result and os.path.exists(result):
                    self.repaired_file = result
                    self.status_label.config(text="修复完成！", fg=self.colors['success'])
                    self.update_progress(100, "修复完成")
                    
                    # 自动打开文件
                    time.sleep(1)
                    try:
                        os.startfile(result)
                    except:
                        pass
                    
                    self.window.after(2000, self.window.destroy)
                else:
                    self.status_label.config(text="修复失败", fg=self.colors['error'])
                    self.update_progress(100, "修复失败")
                    time.sleep(3)
                    self.window.after(0, self.window.destroy)
                    
            except Exception as e:
                self.status_label.config(text="发生错误", fg=self.colors['error'])
                self.update_progress(100, "发生错误")
                time.sleep(3)
                self.window.after(0, self.window.destroy)
                
        except Exception as e:
            self.status_label.config(text="初始化错误", fg=self.colors['error'])
            self.update_progress(100, "初始化错误")
            time.sleep(3)
            self.window.after(0, self.window.destroy)
    
    def run(self):
        """运行进度窗口"""
        self.window.mainloop()


def main():
    """主函数"""
    if len(sys.argv) > 1:
        # 处理拖拽的文件
        file_path = sys.argv[1]
        
        if not os.path.exists(file_path):
            messagebox.showerror("错误", f"文件不存在: {file_path}")
            return
        
        if not file_path.lower().endswith('.xlsx'):
            messagebox.showerror("错误", "请拖拽.xlsx格式的Excel文件")
            return
        
        # 启动进度窗口
        progress = ProgressWindow(file_path)
        progress.run()
    else:
        # 显示使用说明
        messagebox.showinfo("WPS Excel修复工具", 
                          "使用方法：\n"
                          "请拖拽需要修复的.xlsx文件到本程序图标上\n\n"
                          "程序将自动：\n"
                          "1. 分析文件中的DISPIMG公式\n"
                          "2. 转换为Excel原生图片\n"
                          "3. 显示修复进度\n"
                          "4. 自动打开修复后的文件\n\n"
                          "Version 1.0")


if __name__ == "__main__":
    main()
